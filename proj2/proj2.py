import streamlit as st
import streamlit.components.v1 as stc
import os
import pandas as pd
import math
import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font
import pyautogui
from datetime import datetime
os.system('cls')

start_time = datetime.now()

# function to assign octant
def octant_assign(x,y,z):
	try:
		oct = 0 # initialization
		if(x>=0 and y>=0):
			oct = 1
		elif(x<0 and y>=0):
			oct = 2
		elif(x<0 and y<0):
			oct = 3
		elif(x>=0 and y<0):
			oct = 4

		if(z<0): # reduce number of if conditions
			oct *= -1

		return oct # return integer value

	#except block in case an error occurs anywhere in the above function
	except:
		print('Error in function: octant_assign')
		exit()

#function for octant count
def octant_count(df, range_total, octant, len_df):
	try:
		#mod ranges
		r=0
		while(r < range_total):
			df.at[r+1,'Octant ID']= str(mod*(r) ) + "-" + str( mod*(r+1)-1) 
			r += 1
		
		#for last index range
		df.at[r+1, 'Octant ID']= str(mod*(r)) + "-" + str(len_df - 1)

		#count values for each octant
		for i in octant:
			df[str(i)] = ''
			df.at[0, str(i)] = df["Octant"].value_counts()[i]
			r = 0
			p = 0
			while(r < range_total):
				# if(i not in df["Octant"][p:p + mod].value_counts().index):
				#     df.at[r+1, str(i)] = 0
				# else:
				df.at[r+1, str(i)] = df["Octant"][p:p + mod].value_counts()[i]
				p += mod
				r += 1

				if(p+mod > len_df):
						if(i not in df["Octant"][p:len_df].value_counts().index):
							df.at[r+1, str(i)] = 0
						else:
							# for last index range
							df.at[r+1, str(i)] = df["Octant"][p:len_df].value_counts()[i]

	#except block in case an error occurs anywhere in the above function
	except:
		print('Error in function: octant_count')
		exit()

# function to count each rank and rank1 octant ID and Name
def octant_frequency(df, range_total, octant_name_id_mapping, octant):
	try:
		c = -1 #counter variable
		#while loop only for overall count row
		while(c < 0):
			oct = {} # empty dictionary to store octant with corresponding count value
			t = 0
			for i in octant:
				oct[i] = df.iloc[0, 14+t]
				t += 1 # increment for next octant number

			oct = dict(sorted(oct.items(), key=lambda item: item[1], reverse=True)) # decreasingly sorted using lambda function

			t = 1
			for i in oct:
				df.at[0,'Rank Octant ' + str(i)] = t #since oct is sorted, rank will be assigned in corresponding order
				if(t == 1): # if rank is 1 then fill column 30 and 31
					df.at[0,'Rank1 Octant ID'] = i 
					df.at[0,'Rank1 Octant Name'] = octant_name_id_mapping[str(i)] #Name for corresponding key
				t += 1
			c +=1 #new counter value => 0 hence shifts to else condition

		#else condition for mod count rows
		else:
			for i in range(range_total+1): # number of rows for mod count
				octMod = {} # empty dictionary to store octant with corresponding count value
				t = 0
				for j in octant: 
					octMod[j] = df.iloc[i+1, 14+t] #store key value pair for each iteration of mod count
					t += 1 # increment for next octant number

				octMod = dict(sorted(octMod.items(), key=lambda item: item[1], reverse=True)) # decreasingly sorted using lambda function

				t = 1
				for k in octMod: #since octMod is sorted, rank will be assigned in corresponding order
					df.at[i+1,'Rank Octant ' + str(k)] = t
					if(t == 1): # if rank is 1 then fill column 30 and 31
						df.at[i+1,'Rank1 Octant ID'] = k
						df.at[i+1,'Rank1 Octant Name'] = octant_name_id_mapping[str(k)] #Name for corresponding key 
					t += 1

	#except block in case an error occurs anywhere in the above function
	except:
			print('Error while executing octant_frequency function')
			exit()

#function to map octant names according to rank1 and create count of rank1 mod values
def octant_rank(df, octant, range_total):
	try:
		i = 0
		for j in octant:
			df.insert(22+i, 'Rank Octant '+ str(j), '', True)
			i += 1

	# insert column 30 and 31
		df.insert(30, 'Rank1 Octant ID', "", True) 
		df.insert(31, 'Rank1 Octant Name', "", True) 

		# given dictionary with names corresponding to each octant
		octant_name_id_mapping = {'1' : "Internal outward interaction",'-1': "External outward interaction",'2' : "External Ejection",'-2': "Internal Ejection",'3' : "External inward interaction",'-3': "Internal inward interaction",'4' : "Internal sweep",'-4': "External sweep"}

		# Function call for assigning ranks & rank1 Octant ID and Octant Name
		octant_frequency(df, range_total, octant_name_id_mapping, octant)
		
		# Headings in corresponding cells for count of Rank 1 for mod Values
		index = range_total + 3 # To ensure 3 row spacing after mod table
		df.iloc[index+2, 28] = 'Octant ID'
		df.iloc[index+2, 29] = 'Octant Name'
		df.iloc[index+2, 30] = 'Count of Rank 1 mod values'
		
		# Create a list and store rank1 count for each octant ID
		countRank1 = [0,0,0,0,0,0,0,0] #list of size 8 initialized to 0
		for i in range (0, range_total+1): # iterating over number of ranges for corresponding mod value
			#evaluates index of Rank1 octant ID from octant
			#and increases count of that index in the list countRank1
			countRank1[octant.index(df.at[i+1,'Rank1 Octant ID'])] += 1 
			
		# Insert the values from list countRank1
		i = 0
		for octant, name in octant_name_id_mapping.items(): # iterating over values of dictionary octant_name_id_mapping
			df.iloc[index+3+i, 28] = octant # insert octant ID
			df.iloc[index+3+i, 29] = name # insert corresponding octant name
			df.iloc[index+3+i, 30] = countRank1[i] # insert count of Rank1 for corresponding Octant ID
			i+=1 #increment for next octant ID

	# except block in case an error occurs anywhere in the above function
	except:
		print('Error in function: octant_rank')
		exit()

#function to create empty table of transition count
def table_transition(df, range_left, range_right, index, octant):

	try:

		#empty heading
		table_name = ""
		#Assign Table name
		if (range_left < 0): #initialized to -1 later
			df.insert(32, '', "", True)
			df.insert(33, '', "", True)
			df.insert(34, 'Overall Transition Count', "", True)
			df.insert(35, '', "", True)
			table_name += "Overall Transition Count"
		else:
			table_name += "Mod Transition Count"
			df.iloc[index+1,34] = str(range_left) + '-' + str(range_right) 

		df.iloc[index,34] = table_name
		df.iloc[index+1,35] = "To"
		index += 2
		df.iloc[index,34] = "Octant #"
		df.iloc[index+1,33] = "From" 

		#assign octant value as row and col headings
		tempVar = 0 #temporary variable initialized to 0
		for i in octant: #iterating in list octant
			if(range_left < 0):
				df.insert(36+tempVar, '', "", True)

			df.iloc[index+1+tempVar,34] = i #row headings
			df.iloc[index,35+tempVar] = i #column headings
			tempVar += 1

		index += 1

		#initialize zero in all cells of the table since str + int concatenation is not allowed while increasing count value
		for row1 in range(0,8):
			for col1 in range(0,8):
				df.iloc[index+row1,35+col1] = 0

	#except block in case an error occurs anywhere in the above function
	except:
		print('Error in function: table_transition')
		exit()

#function to count transitions
def value_transition(df, range_left, range_right, len_df, index):

	try:
		if(range_right != len_df):
			range_right += 1

		#assign octant value of two successive data points to var1 and var2 respectively
		for i in range(range_left,range_right):
			var1 = df.iloc[i,10]
			var2 = df.iloc[i+1,10]

			#if negative octant, we consider modulus of the value
			if (var1 < 0):
				row1 = 2*(abs(var1)-1)+1 #row number = 1,3,5,7 for -1,-2,-3,-4 respectively
			else:
				row1 = 2*(var1-1) #row number = 0,2,4,6 for 1,2,3,4 respectively

			#same in case of columns
			if(var2<0):
				col = 2*(abs(var2)-1)+36
			else:
				col = 2*(var2-1)+35      
			
			df.iloc[row1+index+3,col] += 1 #increase count of octant by 1 for each transition and assign in dedicated cell using .iloc[]

	#except block in case an error occurs anywhere in the above code
	except:
		print('Error in function: value_transition')
		exit()

#function to count longest subsequence length and count
def subsequence_octant(df, len_df, octant, subsequenceTime):
	try:
		time1 = 0 #initial value of starting time    
		count = 1 #initial value of count is 1 since atleast 1 occurance of an octant will be present  i.e. ith element
		for i in range(len_df): #iterating over entire data set
			if(df.iloc[i,10] == df.iloc[i+1,10]): #if two successive octant values are same
				count += 1 #increase count by 1 for (i+1)th element
			
			else:
				num = df.iloc[i,10]
				index = octant.index(num) #index of corrensponding octant number in the list octant [1,-1,2,-2,3,-3,4,-4]
				time2 = df.iloc[i,0]

				if(count == df.iloc[index,45] ): #if longest subsequent count remains same
					df.iloc[index,46] += 1 #increment count of the longest subsequent count i.e. column 46 of corresponding octant number
					subsequenceTime[num].append([time1,time2]) #append a list in subsequentTime[] with start and end times

				elif (count > df.iloc[index,45]): #if longest subsequent count is greater than the previous one
					df.iloc[index,45] = count #new value of longest subsequent count in column 45 of corresponding octant number
					df.iloc[index,46] = 1 #reset count to 1 in coloum 46 of corresponding octant number
					subsequenceTime[num].clear() #reset list if new longest subsequence and found
					subsequenceTime[num].append([time1,time2]) #then append the start and end times

				count = 1
				time1 = df.iloc[i+1,0]

	#except block in case an error occurs anywhere in the above function
	except:
		print('Error in function: subsequence_octant')
		exit()

#function for creating empty table of longest subsequent count
def octant_longest_subsequence_count(df,len_df, octant, subsequenceTime):
	try:

		# insert column empty 43
		df.insert(43, '', "", True)
		# insert column 44 with heading Octant Number
		df.insert(44, 'Octant ##', "", True)
		# insert column 45 with heading Longest Subsequent Length
		df.insert(45, 'Longest Subsequent Length', "", True)
		# insert column 46 with Count
		df.insert(46, 'Count', "", True)
		# insert empty column 47 
		df.insert(47, '', "", True)
		# insert column 48 with Octant Number
		df.insert(48, 'Octant ###', "", True)
		# insert column 49 with heading Longest Subsequent Length
		df.insert(49, 'Longest Subsequent Length', "", True)
		# insert column 50 with Count
		df.insert(50, 'Count', "", True)

		#insert values 1,-1,2,-2,3,-3,4,-4 in column 12
		temp = 0 #temporary variable
		for i in octant:
			df.iloc[temp,44] = i
			temp += 1

		#initialize the table with 0 values since str + int concatenation is not allowed in python
		for i in range(0,8):
			df.iloc[i,45] = 0
			df.iloc[i,46] = 0

		#function call    
		subsequence_octant(df, len_df-1, octant, subsequenceTime)

	#except block in case an error occurs anywhere in the above function
	except:
		print('Error in function: octant_longest_subsequence_count')
		exit()

#function for longest subsequent count with range            
def octant_longest_subsequence_count_with_range(df, octant, subsequenceTime):
	try:

		#store length of dataframe 
		len_df = df.shape[0]

		#function call
		octant_longest_subsequence_count(df,len_df, octant, subsequenceTime)   

		t = 0
		for i in octant: #iterating over the list octant
			index = octant.index(i) #stores the index of each octant number in octant[]
			df.iloc[t,48] = i #assign column 48 with respective octant number
			#copy longest subsequece and count from previous table (col 45,46,47)
			df.iloc[t,49] = df.iloc[index,45] 
			df.iloc[t,50] = df.iloc[index,46] 
			t += 1 

			#increment t to fill respective cells in next row with name Time, From and To
			df.iloc[t,48] = 'Time'
			df.iloc[t,49] = 'From'
			df.iloc[t,50] = 'To'
			t += 1

			#iterating from 0 to count
			for j in range(0, df.iloc[index,46]):
				#fill start and end times of all longest subsequences in col 49 and 50 respectively
				df.iloc[t,49] = subsequenceTime[i][j][0]
				df.iloc[t,50] = subsequenceTime[i][j][1]
				t +=1

	#except block in case an error occurs anywhere in the above function
	except:
		print('Error in function: octant_longest_subsequence_count_with_range')
		exit()

#function for formatting
def formatter(range_total, output):
	try:
		st = Side(style='thin')
		sr = Side(style=None)
		tb = Border(left=st, right=st, top=st, bottom=st) #defining thin border on all sides
		nb = Border(left=sr, right=sr, top=sr, bottom=sr) #defining no border on all sides
		hl = PatternFill(patternType='solid', fgColor='FFFF00') #defining highlights
		f = Font(bold = False) #unbolding

		#opening new workbook 
		wb = openpyxl.load_workbook(output)
		sheet = wb['Sheet1']

		for i in range(1, 52):
			sheet.cell(row = 1, column = i).font = f #unbolding all headings

		#removing borders from headings
		for i in range(1, 14): 
			sheet.cell(row = 1, column = i).border = nb 

		for i in range(33, 45):
			sheet.cell(row = 1, column = i).border = nb 
			
		sheet.cell(row = 1, column = 48).border = nb 
		sheet.cell(row = 1, column = 52).border = nb 
		
		#thin border on count and rank tables
		for i in range(2, range_total+4):
			for j in range(14,33):
				sheet.cell(row = i, column = j).border = tb

		#thin border on count of rank1 values table
		for i in range(range_total+7, range_total+16):
			for j in range(29,32):
				sheet.cell(row = i, column = j).border = tb
		
		#thin border on overall transition count
		for i in range(3, 12):
			for j in range(35,44):
				maxval1 = sheet.cell(row = i, column = j).border = tb

		#thin border on mod transition count
		index = 0
		for i in range(0,range_total+1):
			for j in range(0,9):
				for k in range (35,44):
					sheet.cell(row = index+17+j, column = k).border = tb
			index += 13
		
		#thin border on longest subsequent count
		for i in range(2,10):
			for j in range(45,48):
				sheet.cell(row = i, column = j).border = tb

		#thin border on longest subsequent count with range
		sum = 0
		for i in range (2,10):
			sum += sheet.cell(row = i, column = 47).value
			for i in range(2, sum+18):
				for j in range(49,52):
						sheet.cell(row = i, column = j).border = tb

		#highlighting rank 1 cells
		for i in range(2,range_total+4):
			for j in range(23, 31):
				if(sheet.cell(row = i, column = j).value == 1):
					sheet.cell(row = i, column = j).fill = hl

		#highlighting max value in overall transition count
		maxval1 = 0
		for i in range(4, 12):
			for j in range(35,44):
				if(maxval1 < sheet.cell(row = i, column = j).value):
					maxval1 = sheet.cell(row = i, column = j).value
			for j in range(35,44):
				if(sheet.cell(row = i, column = j).value == maxval1):
					sheet.cell(row = i, column = j).fill = hl
			maxval1 = 0

		#highlighting max value in mod transition count
		maxval2 = 0
		index = 0
		for i in range(0,range_total+1):
			for j in range(1,9):
				for k in range (36,44):
					if(maxval2 < sheet.cell(row = index+17+j, column = k).value):
						maxval2 = sheet.cell(row = index+17+j, column = k).value        
				for k in range (36,44):
					if(sheet.cell(row = index+17+j, column = k).value == maxval2):
						sheet.cell(row = index+17+j, column = k).fill = hl
				maxval2 = 0
			index += 13

		wb.save(output)

	#except block in case an error occurs anywhere in the above function
	except:
		print('Error in function: formatter')
		exit()

#function for pre-processing
def octant_identification(df, output, mod=5000):

	try:
		#store length of dataframe 
		len_df = df.shape[0]

		# insert U Avg, V Avg & W Avg columns with blank values using .insert()
		df.insert(4, 'U Avg', "", True)
		df.insert(5, 'V Avg', "", True)
		df.insert(6, 'W Avg', "", True)

		# calculate mean values of U,V,W and assigned to index 0 row of U Avg, V Avg, W Avg using .at() rounded to 3 decimal
		df.at[0,'U Avg'] = round(df['U'].mean(),3)
		df.at[0,'V Avg'] = round(df['V'].mean(),3)
		df.at[0,'W Avg'] = round(df['W'].mean(),3)

		# insert U', V' & W' columns 
		df.insert(7, 'U\'=U-U Avg', "", True)
		df.insert(8, 'V\'=V-V Avg', "", True)
		df.insert(9, 'W\'=W-W Avg', "", True)

		# calculate U', V', W' for each row using .subtract() rounded to 3 decimal
		df['U\'=U-U Avg'] = round(df['U'].subtract(df['U Avg'][0]),3)
		df['V\'=V-V Avg'] = round(df['V'].subtract(df['V Avg'][0]),3)
		df['W\'=W-W Avg'] = round(df['W'].subtract(df['W Avg'][0]),3)
		
		# lambda function
		# .apply() to pass a function and apply it to all rows
		df['Octant'] = df.apply(lambda row: octant_assign(row["U'=U-U Avg"], row["V'=V-V Avg"], row["W'=W-W Avg"]), axis = 1)

		# insert column empty 11
		df.insert(11, '', "", True)

		#insert column 12
		df.insert(12, ' ', "", True)
		df.at[0, ' '] = 'Mod ' + str(mod) # display mod as string using str()
		
		# insert column 13
		df.insert(13, 'Octant ID', "", True)
		df.at[0, 'Octant ID'] = "Overall Count"

		num = len_df/mod
		range_total = math.floor(num) # total number of ranges
										# .floor() works as greatest integer function

		#list with values of octant ID
		octant = [1, -1, 2, -2, 3, -3, 4, -4] 

		#function calls for count and rank
		octant_count(df, range_total, octant, len_df)
		octant_rank(df, octant, range_total)

		#Overall transition count table
		table_transition(df,-1,-1,-1,octant)
		value_transition(df, 0, len_df-1, len_df-1, -1)

		index = 13 #index of 1st mod transition count heading

		# mod transition count table
		for i in range (range_total+1): 
			range_left = i*mod #define left bound of range
			range_right = min(len_df-1, ((i+1)*mod)-1) #define range bound of range
			table_transition(df,range_left,range_right,index,octant)
			value_transition(df, range_left,range_right, len_df-1, index)
			index += 13 #increment by 13 for next table heading

		#longest subsequent count function call
		subsequenceTime = [[],[],[],[],[],[],[],[],[]] #list for storing start and end times
		octant_longest_subsequence_count_with_range(df, octant, subsequenceTime)

	except:
		print('Error in function: octant_identification')
		exit()

	try:
		# Write over corresponding output file
		df.to_excel(output, index = False)
	except PermissionError:
		print('Permission Error: Cannot overwrite an opened file')
		exit()
	except:
		print('Error in file: Could not overwrite')
		exit()

	#function call for formatting the output file
	formatter(range_total, output)

	return output

#Streamlit Interface Begins
st.set_page_config(
        page_title="Python Project",
        page_icon=":gem:",
    ) #Setting Favicon and page name

hide_streamlit_style = """
            <style>
            #MainMenu {visibility: hidden;}
            footer {visibility: hidden;}
            </style>
            """
st.markdown(hide_streamlit_style, unsafe_allow_html=True) #Hiding default streamlit stuff

st.caption("By Aarav Arya and Nischal Jain")
st.title("Octant Analysis")
st.write("##")
st.markdown("<div id='linkto_top'></div>", unsafe_allow_html=True) #creating html ID

if "outputs" not in st.session_state:
	st.session_state.outputs = []

def reset():
	st.session_state.outputs.clear()

#Radio Button to choose processing type
st.subheader("Processing Type")
bar = st.radio("Select one option", ("Single File", "Multiple Files", "Batch Processing"), on_change = reset, horizontal=True)
st.write("-"*25)

#for single file option
if bar == "Single File":
	
	st.write("##")
	st.subheader("Input Data")
	data_file = st.file_uploader("File Upload", type=['xlsx'])
	st.write("#")
	#making 2 columns to display text and take number input side by side
	c1, c2 = st.columns((2,10))
	with c1:
		st.write("#")
		st.write("MOD Value")
	with c2:
		mod = st.number_input("Mod Value",min_value = 1, value = 5000, label_visibility="hidden")
	st.write("#")

	if st.button("Process File"): #processing begins on pressing the button
		with st.spinner("Processing..."): #added spinner
			st.write("-"*25)

			if data_file is not None:
				
				pref = "output\\"
				suff = "_octant_analysis_mod_" + str(mod) + ".xlsx"
				output = pref+(data_file.name[:-5])+suff

				df = pd.read_excel(data_file)
				if(len(df) != 0):
					if(mod > len(df)): #mod value should be less than or equal to length of dataframe
						st.error('Incorrect MOD Value', icon="🚨")

					else:
						output = octant_identification(df, output, mod) #function call for tut07

						date_time = datetime.now().strftime("%Y-%m-%d-%H-%M-%S")

						#Download button
						with open(output, 'rb') as file:
							st.subheader("Output Data For : " + str(output[7:-28]) + ".xlsx")
							x = pd.read_excel(output)
							x = x.style.highlight_null(props="color: transparent;") #removing NA cells
							st.dataframe(x) #display output dataframe
							st.download_button(label='Download File : ' + (data_file.name[:-5])+"_"+"mod_"+str(mod)+".xlsx", 
							data = file, 
							file_name = (data_file.name[:-5])+"_"+"mod_"+str(mod)+"_"+date_time+".xlsx",
							mime = "application/octet-stream")
				else:
					st.error('Empty File', icon="🚨")

			else:
				st.error('Input File not found', icon="🚨")
	st.write("#")
	#added a Scroll to Top hyperlink which redirects to html id created above (anchor tag)
	st.markdown("<a href='#linkto_top'>Scroll to Top</a>", unsafe_allow_html=True)

#similar for multiple files
if bar == "Multiple Files":

	st.write("##")
	st.subheader("Input Data")
	#file uploader accepts multiple files
	data_file = st.file_uploader("File Upload", type=['xlsx'], accept_multiple_files=True)
		
	st.write("#")
	c1, c2 = st.columns((2,10))
	with c1:
		st.write("#")
		st.write("MOD Value")
	with c2:
		mod = st.number_input("Mod Value",min_value = 1, value = 5000, label_visibility="hidden")

	st.write("#")
	if(len(data_file) > 1): #number of files should be greater than 1
		if st.button("Process Files "):
			with st.spinner("Processing..."):
				st.write("-"*25)

				st.session_state.outputs.clear() #reseting session state variable
				if data_file is not None:
					for i in data_file:
						
						pref = "output\\"
						suff = "_octant_analysis_mod_" + str(mod) + ".xlsx"
						output = pref+(i.name[:-5])+suff

						df = pd.read_excel(i)
						if(len(df) != 0):
							if(mod > len(df)):
								st.error('Incorrect MOD Value', icon="🚨")
								break
							else:

								output = octant_identification(df, output, mod)

								st.session_state.outputs.append(output)
						else:
							st.error('Empty File : ' + str(i.name), icon="🚨")
				
				else:
					st.error('Input Files not found', icon="🚨")

	else:
		st.write('Please select more than one file')

	suff = "_octant_analysis_mod_" + str(mod) + ".xlsx"
	date_time = datetime.now().strftime("%Y-%m-%d-%H-%M-%S")

	#download button doesnt gets refreshed after downloading 1 file since session state variable is preserved
	for output in st.session_state.outputs:					
		with open(output, 'rb') as file:
			st.subheader("Output Data For : " + str(output[7:-29]) + ".xlsx")
			x = pd.read_excel(file)
			x = x.style.highlight_null(props="color: transparent;")
			if st.button("Display Dataframe : " + str(output[7:-29])+"_"+"mod_"+str(mod)+".xlsx", type="primary"):
				st.dataframe(x)
			st.download_button(label='Download File : ' + str(output[7:-29])+"_"+"mod_"+str(mod)+".xlsx", 
			data = file, 
			file_name = (output[7:-29])+"_"+"mod_"+str(mod)+"_"+date_time+".xlsx",
			mime = "application/octet-stream")
	st.write("#")
	st.markdown("<a href='#linkto_top'>Scroll to Top</a>", unsafe_allow_html=True)
#for batch processing via path
if bar == "Batch Processing":
	
	st.write("##")
	st.subheader("Input Data")
	path = st.text_input("Input Path", value="") #text input path
	st.write("#")
	c1, c2 = st.columns((2,10))
	with c1:
		st.write("#")
		st.write("MOD Value")
	with c2:
		mod = st.number_input("Mod Value",min_value = 1, value = 5000, label_visibility="hidden")
	st.write("#")
	if st.button("Process Files  "):	
		with st.spinner("Processing..."):
			st.write("-"*25)
			st.session_state.outputs.clear()
			if(len(path) == 0): #if no path entered
				st.error('Enter a path', icon="🚨")
			elif not os.path.isdir(path): #if path doesnt exists
				st.error('No Such Path Found', icon="🚨")
			else:
				pref = "output\\"
				suff = "_octant_analysis_mod_" + str(mod) + ".xlsx"
				temp = [] 
				if(len(os.listdir(path)) == 0): #if folder is empty
					st.error('No Files in Path', icon="🚨")
				else:
					for filename in os.listdir(path):
						file = os.path.join(path, filename)
						if(file.endswith(".xlsx")):
							temp.append(file) #append all excel files processed 
							name = str(filename)
							input = pd.read_excel(file)
							df = pd.DataFrame(input)
							if(len(df) != 0):

								output = pref+name[:-5]+suff
								
								df = pd.read_excel(file)
								if(mod > len(df)):
									st.error('Incorrect MOD Value', icon="🚨")
									break
								else:
									output = octant_identification(df, output, mod)

									st.session_state.outputs.append(output)

							else:
								st.error('Empty File : ' + str(filename), icon="🚨")
						else:
							pass
					if(len(temp) == 0): #if no excel files exist in folder
						st.error('No Excel Files Found', icon="🚨")

	suff = "_octant_analysis_mod_" + str(mod) + ".xlsx"
	date_time = datetime.now().strftime("%Y-%m-%d-%H-%M-%S")

	#download button doesnt gets refreshed after downloading 1 file since session state variable is preserved
	for output in st.session_state.outputs:					
		with open(output, 'rb') as file:
			st.subheader("Output Data For : " + str(output[7:-29]) + ".xlsx")
			x = pd.read_excel(file)
			x = x.style.highlight_null(props="color: transparent;")
			if st.button("Display Dataframe : " + str(output[7:-29])+"_"+"mod_"+str(mod)+".xlsx", type="primary"):
				st.dataframe(x)
			st.download_button(label='Download File :  ' + str(output[7:-29])+"_"+"mod_"+str(mod)+".xlsx", 
			data = file, 
			file_name = (output[7:-29])+"_"+"mod_"+str(mod)+"_"+date_time+".xlsx",
			mime = "application/octet-stream")
	st.write("#")
	st.markdown("<a href='#linkto_top'>Scroll to Top</a>", unsafe_allow_html=True)

#reset button to refresh the page
col1, midcol, col2 = st.columns((2,10,1.5))
with(col2):
	if st.button("Reset"):
		pyautogui.hotkey("ctrl", "F5")

from platform import python_version
ver = python_version()

if ver == "3.8.10":
	print("Correct Version Installed")
else:
	print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")

#This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
